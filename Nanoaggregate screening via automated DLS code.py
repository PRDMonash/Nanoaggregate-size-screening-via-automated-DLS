from datetime import datetime
import serial
import serial.tools.list_ports
from time import *
from time import sleep
import os
from SF10 import SF10
import subprocess
import pyautogui
import pygetwindow as gw 
import glob
import os.path
import pandas as pd
import numpy as np
from pandas import ExcelWriter
from pandas import ExcelFile
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.ticker as ticker
from matplotlib import cm
import openpyxl 
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
from shutil import copyfile
from colorama import Fore, Style

# Find the COMs that syringe pumps connected to computer
PortData = serial.tools.list_ports.comports()
print(PortData)
for port in PortData:
    print(f"\033[1;31m{port}\033[0m")
    
# Pumps Assigned 
PumpPolymer = SF10('COM11', 'PumpPolymer')
PumpWater = SF10('COM10', 'PumpWater')
PumpDilution = SF10('COM12','PumpDilution')

#open Kalliope software
#subprocess.Popen("C:\Program Files\Anton Paar\Kalliope\AntonPaar.Calliope.exe")
#For the first time, manually enter parameters, give a name to workbook and save it somewhere

#entering operator name and solvents
operator = input('Please input the name of the operator:>> ')
polymer_solvent = input('Please input the polymer solvent:>> ')
second_solvent = input('Please input the second solvent:>> ')
polymer_stocksolution_conc = input('Please input the concentration of the polymer stock solution (mg/ml):>> ')

# Get the current date and time
current_datetime = datetime.now()
# Format the datetime as a string
time_for_report = current_datetime.strftime("%Y-%m-%d %H-%M-%S")

#saving the input file in reports folder
def save_inputexcel(original_path, new_folder, new_name):
    df_input = pd.read_excel(original_path, sheet_name='Sheet1')  
    new_path = os.path.join(new_folder, new_name)
    df_input.to_excel(new_path, index=False)
    print(Fore.GREEN +  f"Input Excel sheet copied and saved to: {new_path}")
    print(Style.RESET_ALL)
    return new_path

original_excel_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Three Pumps Control\excel\flow parameters setup.xlsx"
new_folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Reports"
new_excel_name = "Input Report_{}_{}.xlsx".format(operator, time_for_report) 
saved_input_path = save_inputexcel(original_excel_path, new_folder_path, new_excel_name)

#creating an output report for the experiment
path_to_report = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Reports"
report_workbook = Workbook()
sheet = report_workbook.active  # Get the active sheet (default sheet)
# Add data to the sheet
sheet["A1"] = "Date :"
sheet["A1"].font = Font(bold=True) #making the text bold
sheet["B1"] = datetime.now().date() # adding the current date
sheet["D1"] = "Operator :"
sheet["D1"].font = Font(bold=True) #making the text bold
sheet["E1"] = operator
sheet["G1"] = "Polymer solvent :"
sheet["G1"].font = Font(bold=True)
sheet["H1"] = polymer_solvent
sheet["J1"] = "Second solvent :"
sheet["J1"].font = Font(bold=True)
sheet["K1"] = second_solvent
sheet["M1"] = "Concentration of polymer stock solution (mg/ml) :"
sheet["M1"].font = Font(bold=True)
sheet["N1"] = polymer_stocksolution_conc

sheet["A3"] = "Input :"
sheet["A3"].font = Font(bold=True)
sheet["P3"] = "Output :"
sheet["P3"].font = Font(bold=True)

output_titles = {"A4" : "Experiment no.", "B4" : "Polymer type", "C4" : "Block 1 DP", "D4" : "Block 2 DP", "E4" : "polymer name", 
           "F4" : "concentration of polymer (mg/ml)", "G4" : "Total flow rate (ml/min)", "H4" : "water mixing ratio %", "I4" : "polymer mixing ratio %", "J4" : "repetition", 
           "K4" : "flow rate of water pump (ml/min)", "L4" :"flow rate of polymer overall (ml/min)", "M4" :"stock polymer concentration (mg/ml)", "N4" : "flow rate of polymer pump (ml/min)", "O4" : "flow rate of dilution pump (ml/min)", "P4" : "Measurement name", "Q4" : "Hydrodynamic diameter (µm)", "R4" : "Polydispersity index", 
           "S4" : "Baseline", "T4" : "Peak volume 1 (µm)", "U4" : "Area volume 1 %", "V4" : "Peak volume 2 (µm)", "W4" : "Area volume 2 %", 
           "X4" : "Peak volume 3 (µm)", "Y4" : "Area volume 3 %", "Z4" : "Peak intensity 1 (µm)", "AA4" : "Area intensity 1 %", "AB4" : "Peak intensity 2 (µm)", 
           "AC4" : "Area intensity 2 %", "AD4" : "Peak intensity 3 (µm)", "AE4" : "Area intensity 3 %", "AF4" : "Peak number 1 (µm)", "AG4" : "Area number 1 %", 
           "AH4" : "Peak number 2 (µm)", "AI4" : "Area number 2 %", "AJ4" : "Peak number 3 (µm)", "AK4" : "Area number 3 % "}
for cell, value in output_titles.items():
    sheet[cell] = value
    sheet[cell].font = Font(bold=True, color="000080")
    
report_name = "Output Report_{}_{}.xlsx".format(operator, time_for_report)  
report_path = os.path.join(path_to_report, report_name)
report_workbook.save(report_path)
print(Fore.GREEN +  f"Output Excel sheet created and saved to: {report_path}")
print(Style.RESET_ALL)

#adding input and output files to the report summary in reports folder
summary_path1 = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Reports\Report summary.xlsx"
summary_sheet = "Sheet1"
input_path = saved_input_path
output_path = report_path
summary_workbook = openpyxl.load_workbook(summary_path1)
summary_sheet = summary_workbook[summary_sheet]
summary_sheet.cell(row=summary_sheet.max_row + 1, column=1, value=current_datetime)
summary_sheet.cell(row=summary_sheet.max_row, column=2, value=operator)
summary_sheet.cell(row=summary_sheet.max_row, column=3, value=input_path)
summary_sheet.cell(row=summary_sheet.max_row, column=4, value=output_path)
summary_workbook.save(summary_path1)
print(Fore.GREEN + f"Input and Output Excel sheet links added and saved in: {summary_path1}")
print(Style.RESET_ALL)
   
# Flow rates, mixing profiles and repetitons input

#Reading excel data for input
EXP_NO = "Experiment no."
POLYMER_TYPE = "polymer type"
POLYMER_NAME = "polymer name"
CONC = "concentration of polymer (mg/ml)"
FLOW_RATE = "total flow rate (ml/min)"
WATER_MIX_RATIO = "water mixing ratio %"
POLYMER_MIX_RATIO = "polymer mixing ratio %"
REP = "repetition"
FLOW_RATE_WATER = "flow rate of water pump (ml/min)"
FLOW_RATE_POLYMER = "flow rate of polymer pump (ml/min)"
FLOW_RATE_DILUTION = "flow rate of dilution pump (ml/min)"

df_excel = pd.read_excel(r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Three Pumps Control\excel\flow parameters setup.xlsx")
#print(df_excel)
numeric_columns = df_excel.select_dtypes(include=[float])
df_excel[numeric_columns.columns] = df_excel[numeric_columns.columns].round(2)
row_names = df_excel.index
row_names_list = df_excel.index.tolist()

polymername = df_excel[POLYMER_NAME]
polymer_concentration = df_excel[CONC]
totalflowrates = df_excel[FLOW_RATE]
water_mixing_ratio= df_excel[WATER_MIX_RATIO]
polymer_mixing_ratio= df_excel[POLYMER_MIX_RATIO]
repetition= df_excel[REP]
waterpump_flowrate = df_excel[FLOW_RATE_WATER]
polymerpump_flowrate = df_excel[FLOW_RATE_POLYMER]
dilutionpump_flowrate = df_excel[FLOW_RATE_DILUTION]

def flow_parameters(a, b, c, d, e, f, g):

    #making a new list including repetitions required for water flow rates   
    newlist_water = [ratew for item, ratew in zip(c, a) for repnumber in range(item)]
                    
    #making a new list including repetitions required for polymer flow rates
    newlist_polymerflow = [ratep for item, ratep in zip(c, b) for repnumber in range(item)]
    newlist_dilutionflow = [rated for item, rated in zip(c, g) for repnumber in range(item)] 
    
    #making a new list including repetitions required for polymer names
    newlist_polymername = [raten for item, raten in zip(c, d) for repnumber in range(item)]

    #making a new list including repetitions required for indexes 
    newlist_index = [ratei for item, ratei in zip(c, e) for repnumber in range(item)]
     
    #making a new list including repetitions required for polymer concentrations
    newlist_polymerconc = [ratec for item, ratec in zip(c, f) for repnumber in range(item)]
 
    for flowrates in zip(newlist_water,newlist_polymerflow, newlist_dilutionflow,newlist_polymername,newlist_polymerconc,newlist_index):
        print (flowrates)  #combines five lists for five iterations at the same time
    for flowrates,(wf,pf,df,pn,pc,pi) in enumerate (zip(newlist_water,newlist_polymerflow, newlist_dilutionflow,newlist_polymername,newlist_polymerconc,newlist_index)):
        print (flowrates,wf,pf,df,pn,pc,pi) #gives an index to the zipped output as in a list
    
        print(Fore.GREEN + "Experiment {} is starting : flow rate of water pump : {} ml/min, polymer pump : {} ml/min, dilution pump : {} ml/min for {} at {} mg/ml polymer concentration" .format(flowrates+1,wf,pf,df,pn,pc))
        print(Style.RESET_ALL)

        #the relationship between concentration and flow rates of polymer and dilution pump c1v1=c2v2
        #polymer_concentration = polymer_stocksolution_conc*(pf/(pf+df))
    
        #printing the input row in the output workbook final row
        load_workbook(report_path) #load the workbook
        sheet = report_workbook.active  
        target_row_number = sheet.max_row + 1
        source_row_index = pi  
        row_to_print = df_excel.iloc[source_row_index]
         
        for col_idx, cell_value in enumerate(row_to_print):
            sheet.cell(row=target_row_number, column=col_idx + 1, value=cell_value)
        report_workbook.save(report_path)

        print(Fore.GREEN + "Water will be passed for 1 minute at 1.5 ml/min flow rate")
        print(Style.RESET_ALL)

        # wash with water for 1 minute 
        PumpWater.start()
        sleep(0.5)
        PumpWater.changeFlowrate(1.5) 
        sleep(60)

        print(Fore.GREEN + f"Water and polymer wash will be passed for 30 seconds at flow rates, water: {wf} ml/min, polymer: {pf} ml/min and diluting neat solvent: {df} ml/min")
        print(Style.RESET_ALL)

        #water and polymer and solvent passes for 30 seconds before actual experiment starts
        PumpPolymer.start()
        sleep(0.5)
        PumpDilution.start()
        sleep(0.5)
        PumpWater.changeFlowrate(wf) , PumpPolymer.changeFlowrate(pf) , PumpDilution.changeFlowrate(df)
        sleep(30)

        # Experiment condition input 
        V_cell = 0.2 #float(input('Please input the Volume of your cell (ml):>> '))
        V_dead = 1 #float(input('\033[1;31mPlease input the dead Volume between mixcromixer and cell (ml):>> \033[0m'))

        #sleeptime = time for which the pumps run flling the cell and tubing with micelles
        sleeptime = (V_cell + V_dead)*60/(wf+pf+df)

        #water, polymer and solvent pumps running for sleeptime

        print(Fore.GREEN + "The three pumps will run for {} seconds to fill the cell with sample; water pump : {} ml/min , polymer pump : {} ml/min and dilution pump : {} ml/min" .format(sleeptime,wf,pf,df))
        print(Style.RESET_ALL)

        PumpWater.changeFlowrate(wf) , PumpPolymer.changeFlowrate(pf) , PumpDilution.changeFlowrate(df) 
        sleep(sleeptime)

        PumpWater.stop(), PumpPolymer.stop() , PumpDilution.stop()

        print(Fore.GREEN + "Experiment {} is finished. DLS analysis will start now".format(flowrates+1))
        print(Style.RESET_ALL)

        #DLS analysis 

        #activate kalliope screen
        sleep(3)
        #x=gw.getAllTitles()
        #print(x)
        hwnd = gw.getWindowsWithTitle('apkw')
        print(hwnd)
        if hwnd != []:
            try:
                hwnd[0].activate()
            except:
                hwnd[0].minimize()
                hwnd[0].maximize()

        #click on copy prameters button
        sleep(4)
        copy_click = pyautogui.locateCenterOnScreen("S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Three Pumps Control\copy1.png", confidence=0.5) 
        print (copy_click)
        pyautogui.moveTo(copy_click,duration=2)
        pyautogui.click(copy_click)

        #erases the current title
        sleep(2)
        pyautogui.hotkey("backspace")

        #types the title including current exp no, polymer name, water flow rate and polymer flow rate
        pyautogui.write("Exp {}_{}_water {}_polymer {}_solvent {}_{} mg_ml".format(flowrates+1,pn,wf,pf,df,pc))

        #click on start button
        sleep(3)
        start_click = pyautogui.locateCenterOnScreen("S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Three Pumps Control\start1.png", confidence=0.5) 
        print (start_click)
        pyautogui.moveTo(start_click,duration=2)
        pyautogui.click(start_click)

        #Monitoring for a new file generation in excel files folder
        path_to_watch = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Excel files"
        print(Fore.GREEN + "The path of the folder where your excel data will be saved is", path_to_watch)
        print(Style.RESET_ALL)
        before = dict ([(f, None) for f in os.listdir (path_to_watch)])
        while 1:
            after = dict ([(f, None) for f in os.listdir (path_to_watch)])
            added = [f for f in after if not f in before]
            if added:
                    print("Added: ", ", ".join (added))
                    break
            else:
                    before = after

        #identifying the most recent excel file in the folder
        folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Excel files"
        file_type = r'\*xlsx'
        files = glob.glob(folder_path + file_type)
        max_file = max(files, key=os.path.getctime)

        print(Fore.GREEN + "The path of the folder where the size distribution plots are saved is", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Plots")
        print(Style.RESET_ALL)

        #creating the dataframe for plotting
        sleep(2)
        df = pd.read_excel(max_file)
        #print(df)
        
        df1=df[["Unnamed: 5","Unnamed: 6","Unnamed: 7","Unnamed: 8"]]
        df1.loc[:,'Unnamed: 5']*=1000
        df1.columns=["Particle diameter","Intensity weighted","Volume weighted","Number weighted"]
        df1=df1.dropna(axis=0,how="any")
        df1.drop([4,6],axis=0,inplace=True)
        df1=df1.astype(float)
        df1["Particle diameter"] = df1["Particle diameter"].map('{:.1f}'.format)
        #print(df1)
        time_for_plots = current_datetime.strftime("%Y-%m-%d %H-%M-%S")
        
        #Intensity Weighted size distribution plot
        df1.plot.bar(x="Particle diameter",y="Intensity weighted")
        plt.ylabel("Intensity Weighted %",fontsize=12)
        plt.xlabel("Particle Size (nm)",fontsize=12)
        plt.title("Intensity Weighted size distribution")
        plt.gca().set_xticks(plt.gca().get_xticks()[::4])
        figure = plt.gcf() # get current figure
        figure.set_size_inches(6,9)
        my_path = os.path.abspath(r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Plots")  
        my_file = 'Exp {}_{}_Intensity weighted_{}.png'.format(flowrates+1,pn,time_for_plots)
        plt.savefig(os.path.join(my_path, my_file))  
        plt.close()
        
        #Volume Weighted size distribution plot
        df1.plot.bar(x="Particle diameter",y="Volume weighted")
        plt.ylabel("Volume Weighted %",fontsize=12)
        plt.xlabel("Particle Size (nm)",fontsize=12)
        plt.title("Volume Weighted size distribution")
        plt.gca().set_xticks(plt.gca().get_xticks()[::4])
        figure = plt.gcf() # get current figure
        figure.set_size_inches(6,9)
        my_path = os.path.abspath(r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Plots") 
        my_file = 'Exp {}_{}_Volume weighted_{}.png'.format(flowrates+1,pn,time_for_plots)
        plt.savefig(os.path.join(my_path, my_file))  
        plt.close()
        
        #Number Weighted size distribution plot
        df1.plot.bar(x="Particle diameter",y="Number weighted")
        plt.ylabel("Number Weighted %",fontsize=12)
        plt.xlabel("Particle Size (nm)",fontsize=12)
        plt.title("Number Weighted size distribution")
        plt.gca().set_xticks(plt.gca().get_xticks()[::4])
        figure = plt.gcf() # get current figure
        figure.set_size_inches(6,9)
        my_path = os.path.abspath(r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Plots") 
        my_file = 'Exp {}_{}_Number weighted_{}.png'.format(flowrates+1,pn,time_for_plots)
        plt.savefig(os.path.join(my_path, my_file))  
        plt.close()
        
        #exporting specific data from exported file to output report excel file
        source_file_path = max_file  # the path to source Excel file
        #df_source = pd.read_excel(max_file)
        source_workbook = openpyxl.load_workbook(source_file_path)
        source_sheet = source_workbook['Measurement 0']

        data_series = [
            source_sheet['B2'].value,
            source_sheet['C7'].value,
            source_sheet['C8'].value,
            source_sheet['C10'].value,
            source_sheet['C15'].value,
            source_sheet['C16'].value,
            source_sheet['C18'].value,
            source_sheet['C19'].value,
            source_sheet['C21'].value,
            source_sheet['C22'].value,
            source_sheet['C24'].value,
            source_sheet['C25'].value,
            source_sheet['C27'].value,
            source_sheet['C28'].value,
            source_sheet['C30'].value,
            source_sheet['C31'].value,
            source_sheet['C33'].value,
            source_sheet['C34'].value,
            source_sheet['C36'].value,
            source_sheet['C37'].value,
            source_sheet['C39'].value,
            source_sheet['C40'].value  
        ]

        openpyxl.load_workbook(report_path)
        sheet = report_workbook.active
        columns_to_append = ['P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK']

        for cellvalue, column in enumerate(columns_to_append):
            cell_address = f"{column}{sheet.max_row}"
            sheet[cell_address] = data_series[cellvalue]

        report_workbook.save(report_path)
        print(Fore.GREEN + f"The data is extracted and saved in {report_path}")
        print(Style.RESET_ALL)

        #1_plotting the scatter plots for intensity average particle size vs polymer mixing ratio for a specific polymer and concentration
        
        df_scatter = pd.read_excel(report_path, skiprows=3, header=0)
        df_scatter2 = df_scatter[["polymer name","Total flow rate (ml/min)","polymer mixing ratio %","concentration of polymer (mg/ml)", "Peak intensity 1 (µm)"]]
        df_scatter2.loc[:,'Peak intensity 1 (µm)']*=1000
        df_scatter2["Peak intensity 1 (µm)"] = df_scatter2["Peak intensity 1 (µm)"].map('{:.1f}'.format)
        #print(df_scatter2)

        unique_nameconcentrations = df_scatter2.groupby(['polymer name','concentration of polymer (mg/ml)']).size().reset_index(name='count')
        print(unique_nameconcentrations)

        # Create separate scatter plots for each unique polymer name and concentration
        for index, row in unique_nameconcentrations.iterrows():
            name = row['polymer name']
            concentration = row['concentration of polymer (mg/ml)']
            filtered_df = df_scatter2[(df_scatter2['polymer name'] == name) & (df_scatter2['concentration of polymer (mg/ml)'] == concentration)]
            filtered_df['Peak intensity 1 (µm)'] = pd.to_numeric(filtered_df['Peak intensity 1 (µm)'])
            
            # Create a scatter plot with polymer ratio on the x-axis, size on the y-axis, and total flow rate as the legend
            plt.figure(figsize=(10, 8))
            for flowrate, group in filtered_df.groupby('Total flow rate (ml/min)'):
                plt.scatter(group['polymer mixing ratio %'], group['Peak intensity 1 (µm)'], label=f'{flowrate} (ml/min)', s=70)

            # Customize the plot
            plt.xlabel('Polymer Mixing Ratio %', fontsize = 12, fontweight='bold')
            plt.ylabel('Particle Size (nm)', fontsize = 12, fontweight='bold')
            plt.title(f'Particle size vs polymer mixing ratio scatter plot for\n{name} at {concentration} mg/ml concentration', fontsize = 16, fontweight='bold')
            plt.legend(title='Total Flow rate')
            plt.xticks(fontweight='bold')
            plt.yticks(fontweight='bold')

            my_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
            file_name = 'Particle size vs polymer mixing ratio stacked scatter plot for {} at {} mg_ml.png'.format(name, concentration)
            plt.savefig(os.path.join(my_path, file_name)) 
            plt.close()
        print(Fore.GREEN + "The scatter plot 1 is drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
        print(Style.RESET_ALL)

        #2_plotting the scatter plots for intensity average particle size vs total flow rate for a specific polymer and concentration

        for index, row in unique_nameconcentrations.iterrows():
            name = row['polymer name']
            concentration = row['concentration of polymer (mg/ml)']
            filtered_df = df_scatter2[(df_scatter2['polymer name'] == name) & (df_scatter2['concentration of polymer (mg/ml)'] == concentration)]
            filtered_df['Peak intensity 1 (µm)'] = pd.to_numeric(filtered_df['Peak intensity 1 (µm)'])

            # Create a scatter plot with total flow rate on the x-axis, size on the y-axis, and polymer mix ratio as the legend
            plt.figure(figsize=(10, 8))
            for polymermixratio, group in filtered_df.groupby('polymer mixing ratio %'):
                plt.scatter(group['Total flow rate (ml/min)'], group['Peak intensity 1 (µm)'], label=f'{polymermixratio} %', s=70)

            # Customize the plot
            plt.xlabel('Total flow rate (ml/min)', fontsize = 12, fontweight='bold')
            plt.ylabel('Particle Size (nm)', fontsize = 12, fontweight='bold')
            plt.title(f'Particle size vs total flow rate scatter plot for\n{name} at {concentration} mg/ml concentration', fontsize = 14, fontweight='bold')
            plt.legend(title='Polymer mixing ratio')
            plt.xticks(fontweight='bold')
            plt.yticks(fontweight='bold')

            my_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
            file_name2 = 'Particle size vs total flow rate stacked scatter plot for {} at {} mg_ml.png'.format(name, concentration)
            plt.savefig(os.path.join(my_path, file_name2))
            plt.close()  
        print(Fore.GREEN + "The scatter plot 2 is drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
        print(Style.RESET_ALL)
            
        #3_plotting the scatter plots for intensity average particle size vs polymer mixing ratio for a specific polymer and concentration at a specific flow rate 

        unique_nameflowrate = df_scatter2.groupby(['polymer name','Total flow rate (ml/min)','concentration of polymer (mg/ml)']).size().reset_index(name='count')
        print(unique_nameflowrate)

        # Create separate scatter plots for each unique combination
        for index, row in unique_nameflowrate.iterrows():
            name = row['polymer name']
            flowrate = row['Total flow rate (ml/min)']
            concentration = row['concentration of polymer (mg/ml)']
            filtered_df2 = df_scatter2[(df_scatter2['polymer name'] == name) & (df_scatter2['Total flow rate (ml/min)'] == flowrate) & (df_scatter2['concentration of polymer (mg/ml)'] == concentration)]
            filtered_df2['Peak intensity 1 (µm)'] = pd.to_numeric(filtered_df2['Peak intensity 1 (µm)'])

            plt.figure(figsize=(10, 8))
            plt.scatter(filtered_df2['polymer mixing ratio %'], filtered_df2['Peak intensity 1 (µm)'], s=70)
            
            plt.xlabel('Polymer Mixing Ratio %', fontsize = 12, fontweight='bold')
            plt.ylabel('Particle Size (nm)', fontsize = 12, fontweight='bold')
            plt.title(f'Particle size vs polymer mixing ratio scatter plot for\n{concentration} mg/ml {name} at {flowrate} ml/min flow rate', fontsize = 14, fontweight='bold')
            #plt.grid(True)
            plt.xticks(fontweight='bold')
            plt.yticks(fontweight='bold')
            
            folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
            file_name3 = f'Particle size vs polymer mixing ratio scatter plot for {concentration} mg_ml {name} at {flowrate} ml_min.png'
            plt.savefig(os.path.join(folder_path, file_name3))
            plt.close()
        print(Fore.GREEN + "The scatter plot 3 is drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
        print(Style.RESET_ALL)
            
        #4_plotting the scatter plots for intensity average particle size vs total flow rate for a specific polymer and concentration at a specific mixing profile

        unique_namemixingprofile = df_scatter2.groupby(['polymer name','polymer mixing ratio %','concentration of polymer (mg/ml)']).size().reset_index(name='count')
        print(unique_namemixingprofile)

        # Create separate scatter plots for each unique combination
        for index, row in unique_namemixingprofile.iterrows():
            name = row['polymer name']
            mixratio = row['polymer mixing ratio %']
            concentration = row['concentration of polymer (mg/ml)']
            filtered_df3 = df_scatter2[(df_scatter2['polymer name'] == name) & (df_scatter2['polymer mixing ratio %'] == mixratio) & (df_scatter2['concentration of polymer (mg/ml)'] == concentration)]
            filtered_df3['Peak intensity 1 (µm)'] = pd.to_numeric(filtered_df3['Peak intensity 1 (µm)'])

            plt.figure(figsize=(10, 8))
            plt.scatter(filtered_df3['Total flow rate (ml/min)'], filtered_df3['Peak intensity 1 (µm)'], s=70)
            
            plt.xlabel('Total flow rate (ml/min)', fontsize = 12, fontweight='bold')
            plt.ylabel('Particle Size (nm)', fontsize = 12, fontweight='bold')
            plt.title(f'Particle size vs total flow rate scatter plot for\n{concentration} mg/ml {name} at {mixratio} % polymer mixing ratio', fontsize = 14, fontweight='bold')
            #plt.grid(True)
            plt.xticks(fontweight='bold')
            plt.yticks(fontweight='bold')
            
            folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
            file_name4 = f'Particle size vs total flow rate scatter plot for {concentration} mg_ml {name} at {mixratio} percent polymer mixing ratio.png'
            plt.savefig(os.path.join(folder_path, file_name4))
            plt.close()
        print(Fore.GREEN + "The scatter plot 4 is drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
        print(Style.RESET_ALL)

        #5_3D surface plot for intensity average particle size vs total flow rate and mixing profile for a specific polymer concentration
        df_scatter = pd.read_excel(report_path, skiprows=3, header=0)
        df_scatter1 = df_scatter[["polymer name","Total flow rate (ml/min)","polymer mixing ratio %","concentration of polymer (mg/ml)", "Peak intensity 1 (µm)"]]
        df_scatter1.loc[:,'Peak intensity 1 (µm)']*=1000
        df_scatter1["Peak intensity 1 (µm)"] = df_scatter1["Peak intensity 1 (µm)"].map('{:.0f}'.format)

        unique_nameconcentration = df_scatter1.groupby(['polymer name','concentration of polymer (mg/ml)']).size().reset_index(name='count')
        print(unique_nameconcentration)

        # Create separate scatter3D plots for each unique polymer name and concentration
        for index, row in unique_nameconcentration.iterrows():
            name = row['polymer name']
            concentration = row['concentration of polymer (mg/ml)']
            filtered_df4 = df_scatter1[(df_scatter1['polymer name'] == name) & (df_scatter1['concentration of polymer (mg/ml)'] == concentration)]
            filtered_df4['Peak intensity 1 (µm)'] = pd.to_numeric(filtered_df4['Peak intensity 1 (µm)'])
            
            # Create a meshgrid for X, Y, and Z
            X, Y = np.meshgrid(np.unique(filtered_df4['Total flow rate (ml/min)']), np.unique(filtered_df4['polymer mixing ratio %']))
            Z = np.zeros_like(X)
            
            # Populate the Z values based on the corresponding (flowrate, mixing_profile) pair
            for i in range(len(filtered_df4)):
                flowrate_value = filtered_df4['Total flow rate (ml/min)'].iloc[i]
                mixing_profile_value = filtered_df4['polymer mixing ratio %'].iloc[i]
                size_value = filtered_df4['Peak intensity 1 (µm)'].iloc[i]

                idx = np.where((X == flowrate_value) & (Y == mixing_profile_value))
                Z[idx] = size_value
                
            # Create a 3D plot
            fig = plt.figure(figsize=(12, 10))
            ax = fig.add_subplot(111, projection='3d')
            
            # Plot the surface  
            surf = ax.plot_surface(X, Y, Z, cmap='hsv', alpha=0.7, linewidth=0, antialiased=False)
            # Create a separate Axes for the color bar
            cax = fig.add_axes([0.92, 0.15, 0.04, 0.65])
            
            # Add color bar with explicit cax parameter
            mappable = cm.ScalarMappable(cmap=cm.hsv)
            mappable.set_array(Z)
            cbar = plt.colorbar(mappable, cax=cax)
            cbar.set_label('Particle size (nm)', fontsize = 14, fontweight='bold', rotation=270, labelpad=15)
            cbar.ax.tick_params(labelsize=14)
            
            ax.set_xlabel('Total flow rate (ml/min)', labelpad = 10, fontweight='bold', fontsize = 14, alpha=0.9)
            ax.set_ylabel('Polymer Mixing Ratio %', labelpad = 10, fontweight='bold', fontsize = 14, alpha=0.9)
            ax.set_zlabel('Particle Size (nm)', labelpad = 10, fontweight='bold', fontsize = 14, alpha=0.9)
            ax.set_title(f'3D surface plot for Particle size variation over total flowrate and polymer mixing Profile\nfor {concentration} mg/ml {name}', fontsize = 16, fontweight='bold', pad = 20, alpha=0.9) 
            ax.tick_params(axis = 'both', which = 'major', labelsize = 14, width=2, length=5)
            ax.tick_params(axis = 'both', which = 'minor', labelsize = 14, width=1.5, length=3) 
            cbar.outline.set_linewidth(2)
            ax.xaxis.pane.fill = False
            ax.yaxis.pane.fill = False
            ax.zaxis.pane.fill = False
            
            folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
            file_name5 = f'3D surface plot for particle size vs flow rate vs polymer ratio for {concentration} mg_ml {name}.png'  
            plt.savefig(os.path.join(folder_path, file_name5), dpi=300, transparent=True, bbox_inches='tight', pad_inches=0.1)
            plt.close()

        #6_3D scatter plot for intensity average particle size vs total flow rate and mixing profile for a specific polymer concentration
            for i in range(len(filtered_df4)):
                flowrate_value = filtered_df4['Total flow rate (ml/min)'].iloc[i]
                mixing_profile_value = filtered_df4['polymer mixing ratio %'].iloc[i]
                size_value = filtered_df4['Peak intensity 1 (µm)'].iloc[i]

                idx = np.where((X == flowrate_value) & (Y == mixing_profile_value))
                Z[idx] = size_value
                
            # Create a 3D scatter plot
            fig = plt.figure(figsize=(10, 8))
            ax = plt.axes(projection='3d')

            # Scatter plots for each group
            my_cmap = plt.get_cmap('hsv')
            sctt = ax.scatter3D(X, Y, Z, alpha = 0.8, c = (X + Y + Z), cmap = my_cmap, s=70)
            ax.set_xlabel('Total flow rate (ml/min)', labelpad = 10, fontweight='bold', fontsize = 13, alpha=0.9)
            ax.set_ylabel('Polymer Mixing Ratio %', labelpad = 10, fontweight='bold', fontsize = 13, alpha=0.9)
            ax.set_zlabel('Particle Size (nm)', labelpad = 10, fontweight='bold', fontsize = 13, alpha=0.9)
            ax.set_title(f'3D Scatter plot for Particle size variation over total flowrate and polymer mixing Profile\nfor {concentration} mg/ml {name}', fontsize = 16, fontweight='bold', pad = 20, alpha=0.9) 
            ax.tick_params(axis = 'both', which = 'major', labelsize = 14, width=2, length=5)
            ax.tick_params(axis = 'both', which = 'minor', labelsize = 14, width=1.5, length=3) 
            ax.grid(color ='grey', linestyle ='-.', linewidth = 0.3, alpha = 0.2) 
            ax.xaxis.pane.fill = False
            ax.yaxis.pane.fill = False
            ax.zaxis.pane.fill = False
            
            cax = fig.add_axes([0.92, 0.15, 0.04, 0.65])
            mappable = cm.ScalarMappable(cmap=cm.hsv)
            mappable.set_array(Z)
            cbar = plt.colorbar(mappable, cax=cax)
            cbar.set_label('Particle size (nm)', fontsize = 14, fontweight='bold', rotation=270, labelpad=15)
            cbar.ax.tick_params(labelsize=14)
            
            folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
            file_name6 = f'3D scatter plot for particle size vs flow rate vs polymer ratio for {concentration} mg_ml {name}.png'  
            plt.savefig(os.path.join(folder_path, file_name6), dpi=300, transparent=True, bbox_inches='tight', pad_inches=0.1)
            plt.close()
        print(Fore.GREEN + "The 3D plot 1 and 2 are drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
        print(Style.RESET_ALL)

        #7_3D scatter plot for intensity average particle size vs total flow rate and mixing profile for a polymer with concentration as the legend

        unique_name = df_scatter1.groupby(['polymer name']).size().reset_index(name='count')
        print(unique_name)

        # Create separate scatter plots for each unique polymer name and concentration
        for index, row in unique_name.iterrows():
            name = row['polymer name']
            filtered_df5 = df_scatter1[(df_scatter1['polymer name'] == name)] 
            filtered_df5['Peak intensity 1 (µm)'] = pd.to_numeric(filtered_df5['Peak intensity 1 (µm)'])
            
            fig = plt.figure(figsize=(12, 10))
            ax = fig.add_subplot(111, projection='3d')
            
            unique_concentrations = filtered_df5['concentration of polymer (mg/ml)'].unique()

            for i, conc in enumerate(unique_concentrations):
                # Filter data for each concentration
                subset = filtered_df5[filtered_df5['concentration of polymer (mg/ml)'] == conc]

            # Plot each subset with a unique color and label
                ax.scatter(subset['Total flow rate (ml/min)'], subset['polymer mixing ratio %'], subset['Peak intensity 1 (µm)'], label=f'{conc} mg/ml', s=70)
            ax.set_xlabel('Total flow rate (ml/min)', labelpad = 10, fontweight='bold', fontsize = 13, alpha=0.9)
            ax.set_ylabel('Polymer Mixing Ratio %', labelpad = 10, fontweight='bold', fontsize = 13, alpha=0.9)
            ax.set_zlabel('Particle Size (nm)', labelpad = 10, fontweight='bold', fontsize = 13, alpha=0.9)
            ax.set_title(f'3D Scatter plot for Particle size variation over total flowrate and polymer mixing Profile\nfor {name}', fontsize = 16, fontweight='bold', pad = 20, alpha=0.9) 
            ax.tick_params(axis = 'both', which = 'major', labelsize = 14, width=2, length=5)
            ax.tick_params(axis = 'both', which = 'minor', labelsize = 14, width=1.5, length=3) 
            ax.grid(color ='grey', linestyle ='-.', linewidth = 0.3, alpha = 0.2) 
            ax.xaxis.pane.fill = False
            ax.yaxis.pane.fill = False
            ax.zaxis.pane.fill = False
            plt.legend(title='Polymer concentration')
            
            folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
            file_name7 = f'3D scatter plot for particle size vs flow rate vs polymer ratio for {name}.png'  
            plt.savefig(os.path.join(folder_path, file_name7))
            plt.close()
        print(Fore.GREEN + "The 3D plot 3 is drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
        print(Style.RESET_ALL)

        print(Fore.GREEN + "All the 3D plots and 2D scatter plots are drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
        print(Style.RESET_ALL)

flow_parameters(list(df_excel[FLOW_RATE_WATER]), list(df_excel[FLOW_RATE_POLYMER]), list(df_excel[REP]), list(df_excel[POLYMER_NAME]), row_names_list, list(df_excel[CONC]), list(df_excel[FLOW_RATE_DILUTION]))

print(Fore.GREEN + "Main experiment is over. Outlier detection will start now. Please check 1_excel files and 2_reports folder in common files and 3_plots, 4_scatter and 3d plots folders in Lakshani folder for data")
print(Style.RESET_ALL)

# finding outliers at the end

df_errordetection = pd.read_excel(report_path, skiprows=3, header=0)
df_errordetection.loc[:,'Peak intensity 1 (µm)']*=1000 

def detect_outliers(data, threshold_percent):
    median = np.median(data) # Calculate the median of the data
    threshold = median * threshold_percent / 100.0 # Calculate the threshold based on the median and threshold_percent
    deviation = (data - median) # Calculate the modified z-score for each data point
    return np.abs(deviation) > threshold

grouped_data = df_errordetection.groupby('Experiment no.') # Group the data by 'Experiment_No'

# Initialize an empty list to store outliers
all_outliers = []

threshold_percent = 20.0

for experiment_no, group_data in grouped_data:
    parameter_values = group_data['Peak intensity 1 (µm)'].tolist()
    
    # Check for outliers within the current experiment
    is_outlier = detect_outliers(parameter_values, threshold_percent)
    
    # Append the detected outliers to the 'all_outliers' list
    experiment_outliers = group_data[is_outlier]
    all_outliers.append(experiment_outliers)

# Concatenate the outlier DataFrames into a single DataFrame
outliers_df = pd.concat(all_outliers)
print(outliers_df)
# 'outliers_df' now contains all the outliers detected within each experiment using a threshold within 20% of the median value

#printing the outliers into a separtae sheet of output report excel 
# Write outliers_df to a new sheet in the existing workbook
with pd.ExcelWriter(report_path, engine='openpyxl', mode='a') as writer:
    outliers_df.to_excel(writer, sheet_name='outliers 1', index=False)

output2workbook = load_workbook(report_path)
output2sheet = output2workbook['outliers 1']
output2workbook.save(report_path)

print(Fore.GREEN + f"The outliers from the data set grouped by exp no. are detected and are saved in {report_path}")
print(Style.RESET_ALL)

# Get the unique rows based on 'Experiment_No' column
unique_outlierrows_df = outliers_df.drop_duplicates(subset='Experiment no.', keep='first')
print(unique_outlierrows_df)

#creating the dataframe to carry out each unique experiment twice
repeated_df = pd.concat([unique_outlierrows_df] * 2, ignore_index=True)
sorted_repeated_df = repeated_df.sort_values(by='Experiment no.')
sorted_repeated_df = sorted_repeated_df.reset_index(drop=True)
print(sorted_repeated_df)

#using the dataframe for experiments by extracting values of columns iterating over each row
for index, row in sorted_repeated_df.iterrows():
    EXP_NUM = row["Experiment no."]
    POL_NAME = row["polymer name"]
    POL_CONC = row["concentration of polymer (mg/ml)"]
    FR = row["Total flow rate (ml/min)"]
    WATER_RATIO = row["water mixing ratio %"]
    POLYMER_RATIO = row["polymer mixing ratio %"]
    REPTN = row["repetition"]
    FR_WATER = row["flow rate of water pump (ml/min)"]
    FR_POLYMER = row["flow rate of polymer pump (ml/min)"]
    FR_DILUTION = row["flow rate of dilution pump (ml/min)"]
    
    print(Fore.GREEN + "Outlier detection experiment {} is starting : flow rate of water pump : {} ml/min, polymer pump : {} ml/min, dilution pump : {} ml/min for {} at {} mg/ml polymer concentration" .format(index+1,FR_WATER,FR_POLYMER,FR_DILUTION,POL_NAME,POL_CONC))
    print(Style.RESET_ALL)

    #printing the input row in the output workbook final row
    report_workbook = load_workbook(report_path)
    sheet = report_workbook["Sheet"]
    target_row_number = sheet.max_row + 1
    source_row_index = index  
    columns_to_print = ["Experiment no.", "Polymer type", "Block 1 DP", "Block 2 DP", "polymer name", "concentration of polymer (mg/ml)", "Total flow rate (ml/min)", "water mixing ratio %", "polymer mixing ratio %", "repetition", "flow rate of water pump (ml/min)", "flow rate of polymer overall (ml/min)", "stock polymer concentration (mg/ml)", "flow rate of polymer pump (ml/min)", "flow rate of dilution pump (ml/min)"]
    row_to_print = sorted_repeated_df.iloc[source_row_index]

    for col_idx, column_name in enumerate(sorted_repeated_df.columns):
        if column_name in columns_to_print:
            cell_value = row[column_name]
            sheet.cell(row=target_row_number, column=col_idx + 1, value=cell_value)
    
    report_workbook.save(report_path)

    print(Fore.GREEN + "Water will be passed for 1 minute at 1.5 ml/min flow rate")
    print(Style.RESET_ALL)

    # wash with water for 1 minute 
    PumpWater.start()
    sleep(0.5)
    PumpWater.changeFlowrate(1.5) 
    sleep(60)

    print(Fore.GREEN + f"Water and polymer wash will be passed for 30 seconds at flow rates, water: {FR_WATER} ml/min, polymer: {FR_POLYMER} ml/min and diluting neat solvent: {FR_DILUTION} ml/min")
    print(Style.RESET_ALL)

    #water and polymer and solvent passes for 30 seconds before actual experiment starts
    PumpPolymer.start()
    sleep(0.5)
    PumpDilution.start()
    sleep(0.5)
    PumpWater.changeFlowrate(FR_WATER) , PumpPolymer.changeFlowrate(FR_POLYMER) , PumpDilution.changeFlowrate(FR_DILUTION)
    sleep(30)

    # Experiment condition input 
    V_cell = 0.2 #float(input('Please input the Volume of your cell (ml):>> '))
    V_dead = 1 #float(input('\033[1;31mPlease input the dead Volume between mixcromixer and cell (ml):>> \033[0m'))

    #sleeptime = time for which the pumps run flling the cell and tubing with micelles
    sleeptime = (V_cell + V_dead)*60/(FR_WATER+FR_POLYMER+FR_DILUTION)

    #water, polymer and solvent pumps running for sleeptime

    print(Fore.GREEN + "The three pumps will run for {} seconds to fill the cell with sample; water pump : {} ml/min , polymer pump : {} ml/min and dilution pump : {} ml/min" .format(sleeptime,FR_WATER,FR_POLYMER,FR_DILUTION))
    print(Style.RESET_ALL)

    PumpWater.changeFlowrate(FR_WATER) , PumpPolymer.changeFlowrate(FR_POLYMER) , PumpDilution.changeFlowrate(FR_DILUTION) 
    sleep(sleeptime)

    PumpWater.stop(), PumpPolymer.stop() , PumpDilution.stop()

    print(Fore.GREEN + "Outlier experiment {} is finished. DLS analysis will start now".format(index+1))
    print(Style.RESET_ALL)

    #DLS analysis    
    sleep(3)
    hwnd = gw.getWindowsWithTitle('apkw')
    print(hwnd)
    if hwnd != []:
        try:
            hwnd[0].activate()
        except:
            hwnd[0].minimize()
            hwnd[0].maximize()

    #click on copy prameters button
    sleep(4)
    copy_click = pyautogui.locateCenterOnScreen("S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Three Pumps Control\copy1.png", confidence=0.5) 
    print (copy_click)
    pyautogui.moveTo(copy_click,duration=2)
    pyautogui.click(copy_click)

    #erases the current title
    sleep(2)
    pyautogui.hotkey("backspace")

    #types the title including current exp no, polymer name, water flow rate and polymer flow rate
    pyautogui.write("OutlierExp {}_Exp no {}_water {}_polymer {}_solvent {}_{} mg_ml".format(index+1,EXP_NUM,FR_WATER,FR_POLYMER,FR_DILUTION,POL_CONC))

    #click on start button
    sleep(3)
    start_click = pyautogui.locateCenterOnScreen("S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Three Pumps Control\start1.png", confidence=0.5) 
    print (start_click)
    pyautogui.moveTo(start_click,duration=2)
    pyautogui.click(start_click)

    #Monitoring for a new file generation in excel files folder
    path_to_watch = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Excel files"
    print(Fore.GREEN + "The path of the folder where your excel data will be saved is", path_to_watch)
    print(Style.RESET_ALL)
    before = dict ([(f, None) for f in os.listdir (path_to_watch)])
    while 1:
        after = dict ([(f, None) for f in os.listdir (path_to_watch)])
        added = [f for f in after if not f in before]
        if added:
                print("Added: ", ", ".join (added))
                break
        else:
                before = after

    #identifying the most recent excel file in the folder
    folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Excel files"
    file_type = r'\*xlsx'
    files = glob.glob(folder_path + file_type)
    max_file = max(files, key=os.path.getctime)

    print(Fore.GREEN + "The path of the folder where the size distribution plots are saved is", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Plots")
    print(Style.RESET_ALL)

    #creating the dataframe for plotting
    sleep(2)
    df = pd.read_excel(max_file)
    #print(df)
    
    df1=df[["Unnamed: 5","Unnamed: 6","Unnamed: 7","Unnamed: 8"]]
    df1.loc[:,'Unnamed: 5']*=1000
    df1.columns=["Particle diameter","Intensity weighted","Volume weighted","Number weighted"]
    df1=df1.dropna(axis=0,how="any")
    df1.drop([4,6],axis=0,inplace=True)
    df1=df1.astype(float)
    df1["Particle diameter"] = df1["Particle diameter"].map('{:.1f}'.format)
    #print(df1)
    time_for_plots = current_datetime.strftime("%Y-%m-%d %H-%M-%S")
    
    #Intensity Weighted size distribution plot
    df1.plot.bar(x="Particle diameter",y="Intensity weighted")
    plt.ylabel("Intensity Weighted %",fontsize=12)
    plt.xlabel("Particle Size (nm)",fontsize=12)
    plt.title("Intensity Weighted size distribution")
    plt.gca().set_xticks(plt.gca().get_xticks()[::4])
    figure = plt.gcf() # get current figure
    figure.set_size_inches(6,9)
    my_path = os.path.abspath(r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Plots")  
    my_file = 'OutlierExp {}_Exp no {}_Intensity weighted_{}.png'.format(index+1,EXP_NUM,time_for_plots)
    plt.savefig(os.path.join(my_path, my_file))  
    plt.close()
    
    #Volume Weighted size distribution plot
    df1.plot.bar(x="Particle diameter",y="Volume weighted")
    plt.ylabel("Volume Weighted %",fontsize=12)
    plt.xlabel("Particle Size (nm)",fontsize=12)
    plt.title("Volume Weighted size distribution")
    plt.gca().set_xticks(plt.gca().get_xticks()[::4])
    figure = plt.gcf() # get current figure
    figure.set_size_inches(6,9)
    my_path = os.path.abspath(r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Plots") 
    my_file = 'OutlierExp {}_Exp no {}_Volume weighted_{}.png'.format(index+1,EXP_NUM,time_for_plots)
    plt.savefig(os.path.join(my_path, my_file))  
    plt.close()
    
    #Number Weighted size distribution plot
    df1.plot.bar(x="Particle diameter",y="Number weighted")
    plt.ylabel("Number Weighted %",fontsize=12)
    plt.xlabel("Particle Size (nm)",fontsize=12)
    plt.title("Number Weighted size distribution")
    plt.gca().set_xticks(plt.gca().get_xticks()[::4])
    figure = plt.gcf() # get current figure
    figure.set_size_inches(6,9)
    my_path = os.path.abspath(r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Plots") 
    my_file = 'OutlierExp {}_Exp no {}_Number weighted_{}.png'.format(index+1,EXP_NUM,time_for_plots)
    plt.savefig(os.path.join(my_path, my_file))  
    plt.close()
    
    #exporting specific data from exported file to output report excel file
    source_file_path = max_file  # the path to source Excel file
    #df_source = pd.read_excel(max_file)
    source_workbook = openpyxl.load_workbook(source_file_path)
    source_sheet = source_workbook['Measurement 0']

    data_series = [
        source_sheet['B2'].value,
        source_sheet['C7'].value,
        source_sheet['C8'].value,
        source_sheet['C10'].value,
        source_sheet['C15'].value,
        source_sheet['C16'].value,
        source_sheet['C18'].value,
        source_sheet['C19'].value,
        source_sheet['C21'].value,
        source_sheet['C22'].value,
        source_sheet['C24'].value,
        source_sheet['C25'].value,
        source_sheet['C27'].value,
        source_sheet['C28'].value,
        source_sheet['C30'].value,
        source_sheet['C31'].value,
        source_sheet['C33'].value,
        source_sheet['C34'].value,
        source_sheet['C36'].value,
        source_sheet['C37'].value,
        source_sheet['C39'].value,
        source_sheet['C40'].value  
    ]

    openpyxl.load_workbook(report_path)
    sheet = report_workbook["Sheet"]
    columns_to_append = ['P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK']

    for cellvalue, column in enumerate(columns_to_append):
        cell_address = f"{column}{sheet.max_row}"
        sheet[cell_address] = data_series[cellvalue]

    report_workbook.save(report_path)
    print(Fore.GREEN + f"The data is extracted and saved in {report_path}")
    print(Style.RESET_ALL)
        
print(Fore.GREEN + "The experiments for outliers are finished. Now the outliers will be analyzed and deleted and plots will be drawn.")
print(Style.RESET_ALL)

df_errordetection2 = pd.read_excel(report_path, sheet_name="Sheet", skiprows=3, header=0)
df_errordetection2.loc[:,'Peak intensity 1 (µm)']*=1000 

def detect_outliers(data, threshold_percent):
    median = np.median(data) # Calculate the median of the data
    threshold = median * threshold_percent / 100.0 # Calculate the threshold based on the median and threshold_percent
    deviation = (data - median) # Calculate the modified z-score for each data point
    return np.abs(deviation) > threshold

grouped_data = df_errordetection2.groupby('Experiment no.') # Group the data by 'Experiment_No'

# Initialize an empty list to store outliers
all_outliers = []

threshold_percent = 20.0

for experiment_no, group_data in grouped_data:
    parameter_values = group_data['Peak intensity 1 (µm)'].tolist()
    
    # Check for outliers within the current experiment
    is_outlier = detect_outliers(parameter_values, threshold_percent)
    
    # Append the detected outliers to the 'all_outliers' list
    experiment_outliers = group_data[is_outlier]
    all_outliers.append(experiment_outliers)

# Concatenate the outlier DataFrames into a single DataFrame
outliers2_df = pd.concat(all_outliers)
print(outliers2_df)

selected_columns = outliers2_df[['Experiment no.', 'Peak intensity 1 (µm)']]
print(selected_columns)

#printing the outliers into output report excel sheet 

# Write outliers2_df to a new sheet in the existing workbook
with pd.ExcelWriter(report_path, engine='openpyxl', mode='a') as writer:
    outliers2_df.to_excel(writer, sheet_name='outliers after repeating exp', index=False)

output2workbook = load_workbook(report_path)
output2sheet = output2workbook['outliers after repeating exp']
output2workbook.save(report_path)

print(Fore.GREEN + f"The outliers after the repeated experiment data are detected and are saved in {report_path}")
print(Style.RESET_ALL)

#dropping outlier rows 
# Reset index to ensure a unique identifier for each row
df_errordetection2 = df_errordetection2.reset_index()
outliers2_df = outliers2_df.reset_index()

def flag_outliers_for_removal(group, outlier_indices):
    # Determine the number of outliers in this group
    num_outliers = sum(group['index'].isin(outlier_indices))
    if num_outliers < 4:
        # If less than 4 outliers, mark them for removal
        group['remove'] = group['index'].isin(outlier_indices)
    else:
        # Otherwise, mark none for removal
        group['remove'] = False
    return group

# Get the indices of the outliers
outlier_indices = outliers2_df['index']

# Apply the function to each group
df_errordetection2 = df_errordetection2.groupby('Experiment no.', group_keys=True).apply(flag_outliers_for_removal, outlier_indices=outlier_indices)
print(df_errordetection2)

df_filtered = df_errordetection2[df_errordetection2['remove'] == False].drop(columns=['remove', 'index'])
print(df_filtered)

# saving this modified output to the same output report created before but on to a different sheet

# Write df_errordetection2 to a new sheet in the existing workbook
with pd.ExcelWriter(report_path, engine='openpyxl', mode='a') as writer:
    df_errordetection2.to_excel(writer, sheet_name='flagged outliers for removal', index=False)

output2workbook = load_workbook(report_path)
output2sheet = output2workbook['flagged outliers for removal']
output2workbook.save(report_path)

# Write df_filtered to a new sheet in the existing workbook
with pd.ExcelWriter(report_path, engine='openpyxl', mode='a') as writer:
    df_filtered.to_excel(writer, sheet_name='output without outliers', index=False)

output2workbook = load_workbook(report_path)
output2sheet = output2workbook['output without outliers']
output2workbook.save(report_path)

print(Fore.GREEN + f"The modified output file with flagged outliers and one after outlier removal are saved in {report_path}")
print(Style.RESET_ALL)

#final plots after removing outliers

#1_plotting the scatter plots for intensity average particle size vs polymer mixing ratio for a specific polymer and concentration
        
df_scatter2 = df_filtered[["polymer name","Total flow rate (ml/min)","polymer mixing ratio %","concentration of polymer (mg/ml)", "Peak intensity 1 (µm)"]]
df_scatter2["Peak intensity 1 (µm)"] = df_scatter2["Peak intensity 1 (µm)"].map('{:.1f}'.format)
#print(df_scatter2)

unique_nameconcentrations = df_scatter2.groupby(['polymer name','concentration of polymer (mg/ml)']).size().reset_index(name='count')
print(unique_nameconcentrations)

# Create separate scatter plots for each unique polymer name and concentration
for index, row in unique_nameconcentrations.iterrows():
    name = row['polymer name']
    concentration = row['concentration of polymer (mg/ml)']
    filtered_df = df_scatter2[(df_scatter2['polymer name'] == name) & (df_scatter2['concentration of polymer (mg/ml)'] == concentration)]
    filtered_df['Peak intensity 1 (µm)'] = pd.to_numeric(filtered_df['Peak intensity 1 (µm)'])
    
    # Create a scatter plot with polymer ratio on the x-axis, size on the y-axis, and total flow rate as the legend
    plt.figure(figsize=(10, 8))
    for flowrate, group in filtered_df.groupby('Total flow rate (ml/min)'):
        plt.scatter(group['polymer mixing ratio %'], group['Peak intensity 1 (µm)'], label=f'{flowrate} (ml/min)', s=70)

    # Customize the plot
    plt.xlabel('Polymer Mixing Ratio %', fontsize = 12, fontweight='bold')
    plt.ylabel('Particle Size (nm)', fontsize = 12, fontweight='bold')
    plt.title(f'Outlier excluded Particle size vs polymer mixing ratio scatter plot for\n{name} at {concentration} mg/ml concentration', fontsize = 16, fontweight='bold')
    plt.legend(title='Total Flow rate')
    plt.xticks(fontweight='bold')
    plt.yticks(fontweight='bold')

    my_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
    file_name = 'Outlier excluded Particle size vs polymer mixing ratio stacked scatter plot for {} at {} mg_ml.png'.format(name, concentration)
    plt.savefig(os.path.join(my_path, file_name)) 
    plt.close()
print(Fore.GREEN + "Outlier excluded scatter plot 1 is drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
print(Style.RESET_ALL)

#2_plotting the scatter plots for intensity average particle size vs total flow rate for a specific polymer and concentration

for index, row in unique_nameconcentrations.iterrows():
    name = row['polymer name']
    concentration = row['concentration of polymer (mg/ml)']
    filtered_df = df_scatter2[(df_scatter2['polymer name'] == name) & (df_scatter2['concentration of polymer (mg/ml)'] == concentration)]
    filtered_df['Peak intensity 1 (µm)'] = pd.to_numeric(filtered_df['Peak intensity 1 (µm)'])

    # Create a scatter plot with total flow rate on the x-axis, size on the y-axis, and polymer mix ratio as the legend
    plt.figure(figsize=(10, 8))
    for polymermixratio, group in filtered_df.groupby('polymer mixing ratio %'):
        plt.scatter(group['Total flow rate (ml/min)'], group['Peak intensity 1 (µm)'], label=f'{polymermixratio} %', s=70)

    # Customize the plot
    plt.xlabel('Total flow rate (ml/min)', fontsize = 12, fontweight='bold')
    plt.ylabel('Particle Size (nm)', fontsize = 12, fontweight='bold')
    plt.title(f'Outlier excluded Particle size vs total flow rate scatter plot for\n{name} at {concentration} mg/ml concentration', fontsize = 14, fontweight='bold')
    plt.legend(title='Polymer mixing ratio')
    plt.xticks(fontweight='bold')
    plt.yticks(fontweight='bold')

    my_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
    file_name2 = 'Outlier excluded Particle size vs total flow rate stacked scatter plot for {} at {} mg_ml.png'.format(name, concentration)
    plt.savefig(os.path.join(my_path, file_name2))
    plt.close()  
print(Fore.GREEN + "Outlier excluded scatter plot 2 is drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
print(Style.RESET_ALL)
    
#3_plotting the scatter plots for intensity average particle size vs polymer mixing ratio for a specific polymer and concentration at a specific flow rate 

unique_nameflowrate = df_scatter2.groupby(['polymer name','Total flow rate (ml/min)','concentration of polymer (mg/ml)']).size().reset_index(name='count')
print(unique_nameflowrate)

# Create separate scatter plots for each unique combination
for index, row in unique_nameflowrate.iterrows():
    name = row['polymer name']
    flowrate = row['Total flow rate (ml/min)']
    concentration = row['concentration of polymer (mg/ml)']
    filtered_df2 = df_scatter2[(df_scatter2['polymer name'] == name) & (df_scatter2['Total flow rate (ml/min)'] == flowrate) & (df_scatter2['concentration of polymer (mg/ml)'] == concentration)]
    filtered_df2['Peak intensity 1 (µm)'] = pd.to_numeric(filtered_df2['Peak intensity 1 (µm)'])

    plt.figure(figsize=(10, 8))
    plt.scatter(filtered_df2['polymer mixing ratio %'], filtered_df2['Peak intensity 1 (µm)'], s=70)
    
    plt.xlabel('Polymer Mixing Ratio %', fontsize = 12, fontweight='bold')
    plt.ylabel('Particle Size (nm)', fontsize = 12, fontweight='bold')
    plt.title(f'Outlier excluded Particle size vs polymer mixing ratio scatter plot for\n{concentration} mg/ml {name} at {flowrate} ml/min flow rate', fontsize = 14, fontweight='bold')
    #plt.grid(True)
    plt.xticks(fontweight='bold')
    plt.yticks(fontweight='bold')
    
    folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
    file_name3 = f'Outlier excluded Particle size vs polymer mixing ratio scatter plot for {concentration} mg_ml {name} at {flowrate} ml_min.png'
    plt.savefig(os.path.join(folder_path, file_name3))
    plt.close()
print(Fore.GREEN + "Outlier excluded scatter plot 3 is drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
print(Style.RESET_ALL)
    
#4_plotting the scatter plots for intensity average particle size vs total flow rate for a specific polymer and concentration at a specific mixing profile

unique_namemixingprofile = df_scatter2.groupby(['polymer name','polymer mixing ratio %','concentration of polymer (mg/ml)']).size().reset_index(name='count')
print(unique_namemixingprofile)

# Create separate scatter plots for each unique combination
for index, row in unique_namemixingprofile.iterrows():
    name = row['polymer name']
    mixratio = row['polymer mixing ratio %']
    concentration = row['concentration of polymer (mg/ml)']
    filtered_df3 = df_scatter2[(df_scatter2['polymer name'] == name) & (df_scatter2['polymer mixing ratio %'] == mixratio) & (df_scatter2['concentration of polymer (mg/ml)'] == concentration)]
    filtered_df3['Peak intensity 1 (µm)'] = pd.to_numeric(filtered_df3['Peak intensity 1 (µm)'])

    plt.figure(figsize=(10, 8))
    plt.scatter(filtered_df3['Total flow rate (ml/min)'], filtered_df3['Peak intensity 1 (µm)'], s=70)
    
    plt.xlabel('Total flow rate (ml/min)', fontsize = 12, fontweight='bold')
    plt.ylabel('Particle Size (nm)', fontsize = 12, fontweight='bold')
    plt.title(f'Outlier excluded Particle size vs total flow rate scatter plot for\n{concentration} mg/ml {name} at {mixratio} % polymer mixing ratio', fontsize = 14, fontweight='bold')
    #plt.grid(True)
    plt.xticks(fontweight='bold')
    plt.yticks(fontweight='bold')
    
    folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
    file_name4 = f'Outlier excluded Particle size vs total flow rate scatter plot for {concentration} mg_ml {name} at {mixratio} percent polymer mixing ratio.png'
    plt.savefig(os.path.join(folder_path, file_name4))
    plt.close()
print(Fore.GREEN + "Outlier excluded scatter plot 4 is drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
print(Style.RESET_ALL)

#5_3D surface plot for intensity average particle size vs total flow rate and mixing profile for a specific polymer concentration

df_scatter1 = df_filtered[["polymer name","Total flow rate (ml/min)","polymer mixing ratio %","concentration of polymer (mg/ml)", "Peak intensity 1 (µm)"]]
df_scatter1["Peak intensity 1 (µm)"] = df_scatter1["Peak intensity 1 (µm)"].map('{:.0f}'.format)

unique_nameconcentration = df_scatter1.groupby(['polymer name','concentration of polymer (mg/ml)']).size().reset_index(name='count')
print(unique_nameconcentration)

# Create separate scatter3D plots for each unique polymer name and concentration
for index, row in unique_nameconcentration.iterrows():
    name = row['polymer name']
    concentration = row['concentration of polymer (mg/ml)']
    filtered_df4 = df_scatter1[(df_scatter1['polymer name'] == name) & (df_scatter1['concentration of polymer (mg/ml)'] == concentration)]
    filtered_df4['Peak intensity 1 (µm)'] = pd.to_numeric(filtered_df4['Peak intensity 1 (µm)'])
    
    # Create a meshgrid for X, Y, and Z
    X, Y = np.meshgrid(np.unique(filtered_df4['Total flow rate (ml/min)']), np.unique(filtered_df4['polymer mixing ratio %']))
    Z = np.zeros_like(X)
    
    # Populate the Z values based on the corresponding (flowrate, mixing_profile) pair
    for i in range(len(filtered_df4)):
        flowrate_value = filtered_df4['Total flow rate (ml/min)'].iloc[i]
        mixing_profile_value = filtered_df4['polymer mixing ratio %'].iloc[i]
        size_value = filtered_df4['Peak intensity 1 (µm)'].iloc[i]

        idx = np.where((X == flowrate_value) & (Y == mixing_profile_value))
        Z[idx] = size_value
        
    # Create a 3D plot
    fig = plt.figure(figsize=(12, 10))
    ax = fig.add_subplot(111, projection='3d')
    
    # Plot the surface  
    surf = ax.plot_surface(X, Y, Z, cmap='hsv', alpha=0.7, linewidth=0, antialiased=False)
    # Create a separate Axes for the color bar
    cax = fig.add_axes([0.92, 0.15, 0.04, 0.65])
    
    # Add color bar with explicit cax parameter
    mappable = cm.ScalarMappable(cmap=cm.hsv)
    mappable.set_array(Z)
    cbar = plt.colorbar(mappable, cax=cax)
    cbar.set_label('Particle size (nm)', fontsize = 14, fontweight='bold', rotation=270, labelpad=15)
    cbar.ax.tick_params(labelsize=14)
    
    ax.set_xlabel('Total flow rate (ml/min)', labelpad = 10, fontweight='bold', fontsize = 14, alpha=0.9)
    ax.set_ylabel('Polymer Mixing Ratio %', labelpad = 10, fontweight='bold', fontsize = 14, alpha=0.9)
    ax.set_zlabel('Particle Size (nm)', labelpad = 10, fontweight='bold', fontsize = 14, alpha=0.9)
    ax.set_title(f'Outlier excluded 3D surface plot for Particle size variation over total flowrate and polymer mixing Profile\nfor {concentration} mg/ml {name}', fontsize = 16, fontweight='bold', pad = 20, alpha=0.9) 
    ax.tick_params(axis = 'both', which = 'major', labelsize = 14, width=2, length=5)
    ax.tick_params(axis = 'both', which = 'minor', labelsize = 14, width=1.5, length=3) 
    cbar.outline.set_linewidth(2)
    ax.xaxis.pane.fill = False
    ax.yaxis.pane.fill = False
    ax.zaxis.pane.fill = False
    
    folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
    file_name5 = f'Outlier excluded 3D surface plot for particle size vs flow rate vs polymer ratio for {concentration} mg_ml {name}.png'  
    plt.savefig(os.path.join(folder_path, file_name5), dpi=300, transparent=True, bbox_inches='tight', pad_inches=0.1)
    plt.close()

#6_3D scatter plot for intensity average particle size vs total flow rate and mixing profile for a specific polymer concentration
    for i in range(len(filtered_df4)):
        flowrate_value = filtered_df4['Total flow rate (ml/min)'].iloc[i]
        mixing_profile_value = filtered_df4['polymer mixing ratio %'].iloc[i]
        size_value = filtered_df4['Peak intensity 1 (µm)'].iloc[i]

        idx = np.where((X == flowrate_value) & (Y == mixing_profile_value))
        Z[idx] = size_value
        
    # Create a 3D scatter plot
    fig = plt.figure(figsize=(10, 8))
    ax = plt.axes(projection='3d')

    # Scatter plots for each group
    my_cmap = plt.get_cmap('hsv')
    sctt = ax.scatter3D(X, Y, Z, alpha = 0.8, c = (X + Y + Z), cmap = my_cmap, s=70)
    ax.set_xlabel('Total flow rate (ml/min)', labelpad = 10, fontweight='bold', fontsize = 13, alpha=0.9)
    ax.set_ylabel('Polymer Mixing Ratio %', labelpad = 10, fontweight='bold', fontsize = 13, alpha=0.9)
    ax.set_zlabel('Particle Size (nm)', labelpad = 10, fontweight='bold', fontsize = 13, alpha=0.9)
    ax.set_title(f'Outlier excluded 3D Scatter plot for Particle size variation over total flowrate and polymer mixing Profile\nfor {concentration} mg/ml {name}', fontsize = 16, fontweight='bold', pad = 20, alpha=0.9) 
    ax.tick_params(axis = 'both', which = 'major', labelsize = 14, width=2, length=5)
    ax.tick_params(axis = 'both', which = 'minor', labelsize = 14, width=1.5, length=3) 
    ax.grid(color ='grey', linestyle ='-.', linewidth = 0.3, alpha = 0.2) 
    ax.xaxis.pane.fill = False
    ax.yaxis.pane.fill = False
    ax.zaxis.pane.fill = False
    
    cax = fig.add_axes([0.92, 0.15, 0.04, 0.65])
    mappable = cm.ScalarMappable(cmap=cm.hsv)
    mappable.set_array(Z)
    cbar = plt.colorbar(mappable, cax=cax)
    cbar.set_label('Particle size (nm)', fontsize = 14, fontweight='bold', rotation=270, labelpad=15)
    cbar.ax.tick_params(labelsize=14)
    
    folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
    file_name6 = f'Outlier excluded 3D scatter plot for particle size vs flow rate vs polymer ratio for {concentration} mg_ml {name}.png'  
    plt.savefig(os.path.join(folder_path, file_name6), dpi=300, transparent=True, bbox_inches='tight', pad_inches=0.1)
    plt.close()
print(Fore.GREEN + "Outlier excluded 3D plot 1 and 2 are drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
print(Style.RESET_ALL)

#7_3D scatter plot for intensity average particle size vs total flow rate and mixing profile for a polymer with concentration as the legend

unique_name = df_scatter1.groupby(['polymer name']).size().reset_index(name='count')
print(unique_name)

# Create separate scatter plots for each unique polymer name and concentration
for index, row in unique_name.iterrows():
    name = row['polymer name']
    filtered_df5 = df_scatter1[(df_scatter1['polymer name'] == name)] 
    filtered_df5['Peak intensity 1 (µm)'] = pd.to_numeric(filtered_df5['Peak intensity 1 (µm)'])
    
    fig = plt.figure(figsize=(12, 10))
    ax = fig.add_subplot(111, projection='3d')
    
    unique_concentrations = filtered_df5['concentration of polymer (mg/ml)'].unique()

    for i, conc in enumerate(unique_concentrations):
        # Filter data for each concentration
        subset = filtered_df5[filtered_df5['concentration of polymer (mg/ml)'] == conc]

    # Plot each subset with a unique color and label
        ax.scatter(subset['Total flow rate (ml/min)'], subset['polymer mixing ratio %'], subset['Peak intensity 1 (µm)'], label=f'{conc} mg/ml', s=70)
    ax.set_xlabel('Total flow rate (ml/min)', labelpad = 10, fontweight='bold', fontsize = 13, alpha=0.9)
    ax.set_ylabel('Polymer Mixing Ratio %', labelpad = 10, fontweight='bold', fontsize = 13, alpha=0.9)
    ax.set_zlabel('Particle Size (nm)', labelpad = 10, fontweight='bold', fontsize = 13, alpha=0.9)
    ax.set_title(f'Outlier excluded 3D Scatter plot\nfor Particle size variation over total flowrate and polymer mixing Profile\nfor {name}', fontsize = 16, fontweight='bold', pad = 20, alpha=0.9) 
    ax.tick_params(axis = 'both', which = 'major', labelsize = 14, width=2, length=5)
    ax.tick_params(axis = 'both', which = 'minor', labelsize = 14, width=1.5, length=3) 
    ax.grid(color ='grey', linestyle ='-.', linewidth = 0.3, alpha = 0.2) 
    ax.xaxis.pane.fill = False
    ax.yaxis.pane.fill = False
    ax.zaxis.pane.fill = False
    plt.legend(title='Polymer concentration')
    
    folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots"
    file_name7 = f'Outlier excluded 3D scatter plot for particle size vs flow rate vs polymer ratio for {name}.png'  
    plt.savefig(os.path.join(folder_path, file_name7))
    plt.close()
print(Fore.GREEN + "Outlier excluded 3D plot 3 is drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
print(Style.RESET_ALL)

print(Fore.GREEN + "All outlier excluded 3D plots and 2D scatter plots are drawn and saved in", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Scatter and 3D plots")
print(Style.RESET_ALL)

print(Fore.GREEN + "Experiment series is over. Please check 1_excel files and 2_reports folder in common files and 3_plots, 4_scatter and 3d plots folders in Lakshani folder for data")
print(Style.RESET_ALL)
